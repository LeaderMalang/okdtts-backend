# inventory/admin.py
from __future__ import annotations
from django.contrib import admin
from .models import Product, Party, Batch, StockMovement, PriceList, PriceListItem
from .forms import PartyForm
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any, Dict
from django import forms
from django.contrib import admin, messages
from django.db import transaction
from django.http import HttpResponse, HttpResponseRedirect
from django.urls import path, reverse
from django.utils.text import slugify
from openpyxl import Workbook, load_workbook
from .models import Product
from inventory.models import Company, Group, Distributor  # adjust path if different


def _norm_header(h: str) -> str:
    return slugify((h or "").strip()).replace("-", "_")


def _to_decimal(v: Any) -> Decimal | None:
    if v is None or v == "":
        return None
    try:
        return Decimal(str(v))
    except (InvalidOperation, ValueError):
        return None


class ProductImportForm(forms.Form):
    file = forms.FileField()


@admin.register(Product)
class ProductAdmin(admin.ModelAdmin):
    list_display = ("name", "company", "rate", "retail_price", "stock", "disable_sale_purchase")
    search_fields = ("name", "barcode")
    list_filter = ("company", "group", "distributor", "disable_sale_purchase")
    actions = ["export_selected_to_excel", "download_import_template"]

    change_list_template = "admin/inventory/product/change_list_with_import.html"

    def get_urls(self):
        urls = super().get_urls()
        my = [
            path("import/", self.admin_site.admin_view(self.import_products_view), name="inventory_product_import"),
            path("export-all/", self.admin_site.admin_view(self.export_all_products_view), name="inventory_product_export_all"),
        ]
        return my + urls

    # ---------- EXPORT ----------
    @admin.action(description="Export selected to Excel (.xlsx)")
    def export_selected_to_excel(self, request, queryset):
        return self._export_queryset_as_xlsx(queryset, filename="products_selected.xlsx")

    def export_all_products_view(self, request):
        qs = self.get_queryset(request)
        return self._export_queryset_as_xlsx(qs, filename="products_all.xlsx")

    def _export_queryset_as_xlsx(self, qs, *, filename: str) -> HttpResponse:
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"

        headers = [
            "Name",
            "Barcode",
            "Company",
            "Group",
            "Distributor",
            "Rate",            # trade_price
            "Retail",          # retail_price
            "Sales Tax Ratio",
            "FED Tax Ratio",
            "Packing",
            "E RATE",          # e_rate
            "Disable Sale/Purchase",
        ]
        ws.append(headers)

        for p in qs.select_related("company", "group", "distributor"):
            ws.append([
                p.name,
                p.barcode or "",
                getattr(p.company, "name", ""),
                getattr(p.group, "name", ""),
                getattr(p.distributor, "name", ""),
                p.trade_price,
                p.retail_price,
                p.sales_tax_ratio,
                p.fed_tax_ratio,
                p.packing or "",
                getattr(p, "e_rate", Decimal("0.00")),
                bool(p.disable_sale_purchase),
            ])

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        resp = HttpResponse(
            bio.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp

    # ---------- IMPORT ----------
    def import_products_view(self, request):
        if request.method == "POST":
            form = ProductImportForm(request.POST, request.FILES)
            if not form.is_valid():
                messages.error(request, "Please upload a valid .xlsx file.")
                return HttpResponseRedirect(reverse("admin:inventory_product_changelist"))

            f = form.cleaned_data["file"]
            try:
                with transaction.atomic():
                    created, updated, disabled = self._import_xlsx(f)
                messages.success(
                    request,
                    f"Import done. Created: {created}, Updated: {updated}, "
                    f"Disabled (E RATE empty/0): {disabled}."
                )
            except Exception as exc:
                messages.error(request, f"Import failed: {exc}")
            return HttpResponseRedirect(reverse("admin:inventory_product_changelist"))

        from django.shortcuts import render
        return render(
            request,
            "admin/inventory/product/import.html",
            {"form": ProductImportForm(), "title": "Import Products from Excel"},
        )

    def _import_xlsx(self, file) -> tuple[int, int, int]:
        """
        Expected headers (exact, but we’ll also tolerate normalized fallbacks):
        - Name, Barcode, Company, Group, Distributor,
          Rate, Retail, Sales Tax Ratio, FED Tax Ratio, Packing, E RATE
        Rules:
        - Match/update by Barcode (preferred); else by (Name + Company).
        - Map: Rate -> trade_price, Retail -> retail_price, E RATE -> e_rate.
        - If E RATE is empty or 0 => disable_sale_purchase=True.
        """
        wb = load_workbook(filename=file, data_only=True)
        ws = wb.active

        # header map
        row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        raw_headers = [str(h or "").strip() for h in row1]
        norm_map: Dict[str, int] = {_norm_header(h): i for i, h in enumerate(raw_headers)}
        exact_map: Dict[str, int] = {h: i for i, h in enumerate(raw_headers)}

        def col_exact(name: str) -> int | None:
            return exact_map.get(name)

        def col_fallback(*names: str) -> int | None:
            # try exact first, then normalized
            for n in names:
                if (i := exact_map.get(n)) is not None:
                    return i
                if (i := norm_map.get(_norm_header(n))) is not None:
                    return i
            return None

        c_name        = col_fallback("Name")
        c_barcode     = col_fallback("Barcode")
        c_company     = col_fallback("Company")
        c_group       = col_fallback("Group")
        c_distributor = col_fallback("Distributor")
        c_rate        = col_fallback("Rate")            # trade_price
        c_retail      = col_fallback("Retail")          # retail_price
        c_sales_tax   = col_fallback("Sales Tax Ratio")
        c_fed_tax     = col_fallback("FED Tax Ratio")
        c_packing     = col_fallback("Packing")
        c_e_rate      = col_fallback("E RATE")          # e_rate

        required = [c_name, c_company, c_group, c_distributor]
        if any(v is None for v in required):
            raise ValueError("Columns required: Name, Company, Group, Distributor")

        created = updated = disabled = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            name = (row[c_name] or "").strip() if c_name is not None and row[c_name] else ""
            if not name:
                continue

            barcode = (row[c_barcode] or "").strip() if c_barcode is not None and row[c_barcode] else ""

            company_name = (row[c_company] or "").strip() if c_company is not None and row[c_company] else ""
            group_name   = (row[c_group] or "").strip() if c_group is not None and row[c_group] else ""
            dist_name    = (row[c_distributor] or "").strip() if c_distributor is not None and row[c_distributor] else ""

            rate   = _to_decimal(row[c_rate]) if c_rate is not None else None
            retail = _to_decimal(row[c_retail]) if c_retail is not None else None
            s_tax  = _to_decimal(row[c_sales_tax]) if c_sales_tax is not None else None
            f_tax  = _to_decimal(row[c_fed_tax]) if c_fed_tax is not None else None
            packing = (row[c_packing] or "").strip() if c_packing is not None and row[c_packing] else ""

            e_rate = _to_decimal(row[c_e_rate]) if c_e_rate is not None else None

            company_obj, _ = Company.objects.get_or_create(name=company_name)
            group_obj, _   = Group.objects.get_or_create(name=group_name)
            dist_obj, _    = Distributor.objects.get_or_create(name=dist_name)

            q = Product.objects.filter(company=company_obj)
            if barcode:
                q = q.filter(barcode=barcode)
            else:
                q = q.filter(name=name)

            to_disable = (e_rate is None or e_rate == Decimal("0"))

            if q.exists():
                p = q.first()
                p.name = name
                p.company = company_obj
                p.group = group_obj
                p.distributor = dist_obj
                if barcode:
                    p.barcode = barcode

                if rate is not None:
                    p.trade_price = rate
                if retail is not None:
                    p.retail_price = retail
                if s_tax is not None:
                    p.sales_tax_ratio = s_tax
                if f_tax is not None:
                    p.fed_tax_ratio = f_tax
                p.packing = packing or None

                if e_rate is not None:
                    p.e_rate = e_rate
                if to_disable:
                    p.disable_sale_purchase = True

                p.save()
                updated += 1
            else:
                p = Product.objects.create(
                    name=name,
                    barcode=barcode or "",
                    company=company_obj,
                    group=group_obj,
                    distributor=dist_obj,
                    trade_price=rate or Decimal("0.00"),
                    retail_price=retail or Decimal("0.00"),
                    sales_tax_ratio=s_tax or Decimal("0.00"),
                    fed_tax_ratio=f_tax or Decimal("0.00"),
                    packing=packing or None,
                    e_rate=e_rate or Decimal("0.00"),
                    disable_sale_purchase=to_disable,
                )
                created += 1

            if p.disable_sale_purchase:
                disabled += 1

        return created, updated, disabled

    @admin.action(description="Download import template (.xlsx)")
    def download_import_template(self, request, queryset):
        wb = Workbook()
        ws = wb.active
        ws.title = "Template"
        ws.append([
            "Name", "Barcode", "Company", "Group", "Distributor",
            "Rate", "Retail", "Sales Tax Ratio", "FED Tax Ratio",
            "Packing", "E RATE"
        ])
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        resp = HttpResponse(
            bio.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = 'attachment; filename="products_import_template.xlsx"'
        return resp

@admin.register(Party)
class PartyAdmin(admin.ModelAdmin):
    form = PartyForm
    list_display = (
        'name',
        'party_type',
        'phone',
        'category',
        'latitude',
        'longitude',
        'price_list',
        "business_image",
        "current_balance"
    )
    search_fields = ('name', 'phone', 'category')
    list_filter = ('party_type',)

@admin.register(Batch)
class BatchAdmin(admin.ModelAdmin):
    list_display = ('product', 'batch_number', 'expiry_date', 'stock', 'rate')
    list_filter = ('product', 'expiry_date')
    search_fields = ('batch_number', 'product__name')

@admin.register(StockMovement)
class StockMovementAdmin(admin.ModelAdmin):
    list_display = ('batch', 'movement_type', 'quantity', 'timestamp', 'reason')
    list_filter = ('movement_type', 'timestamp')
    search_fields = ('batch__batch_number', 'reason')


class PriceListItemInline(admin.TabularInline):
    model = PriceListItem
    extra = 1


@admin.register(PriceList)
class PriceListAdmin(admin.ModelAdmin):
    list_display = ('name',)
    search_fields = ('name',)
    inlines = [PriceListItemInline]
